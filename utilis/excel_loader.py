"""
WatchAI — Excel Profile Loader
Reads criminals/criminals.xlsx and returns a dict of criminal profiles.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import config
from datetime import date

EXPECTED_COLUMNS = [
    "id", "name", "photo", "age", "gender", "crime_type",
    "num_prior_convictions", "weapon_used", "current_status",
    "last_crime_date", "victim_count",
]


def load_criminal_profiles() -> dict:
    """
    Reads criminals.xlsx and returns dict[name → profile].
    Missing optional columns are warned but don't abort loading.
    """
    profiles = {}

    if not os.path.exists(config.CRIMINALS_XLSX):
        print(f"[EXCEL] criminals.xlsx not found: {config.CRIMINALS_XLSX}")
        print("[EXCEL] Run:  uv run python scripts/build_dataset.py")
        return profiles

    wb = openpyxl.load_workbook(config.CRIMINALS_XLSX)
    ws = wb.active
    headers = [str(cell.value).strip().lower() for cell in ws[1]]

    for col in EXPECTED_COLUMNS:
        if col not in headers:
            print(f"[EXCEL] WARNING: Missing column '{col}'")

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        name = row_dict.get("name")
        if not name:
            continue
        name = str(name).strip().lower()

        profile = {
            "id":                    row_dict.get("id"),
            "name":                  name,
            "photo":                 str(row_dict.get("photo") or "").strip(),
            "age":                   _safe_int(row_dict.get("age")),
            "gender":                str(row_dict.get("gender") or "").strip(),
            "crime_type":            str(row_dict.get("crime_type") or "").strip().lower(),
            "num_prior_convictions": _safe_int(row_dict.get("num_prior_convictions"), default=0),
            "weapon_used":           _safe_bool(row_dict.get("weapon_used")),
            "current_status":        str(row_dict.get("current_status") or "").strip().lower(),
            "last_crime_date":       _safe_date(row_dict.get("last_crime_date")),
            "victim_count":          _safe_int(row_dict.get("victim_count"), default=0),
        }
        profiles[name] = profile
        print(f"[EXCEL] {name} | status={profile['current_status']} | crime={profile['crime_type']}")

    print(f"[EXCEL] Total profiles: {len(profiles)}")
    return profiles


# ── Type helpers ──────────────────────────────────────────────────────────────

def _safe_int(value, default=None):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ("true", "yes", "1")
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def _safe_date(value):
    if value is None:
        return None
    if isinstance(value, date):
        return value
    try:
        from datetime import datetime
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None
